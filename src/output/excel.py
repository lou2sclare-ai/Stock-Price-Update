from __future__ import annotations
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

NAVY="0B1F3A"; NAVY2="173A70"; BLUE="2F61D5"; LIGHT="EAF2FF"; PALE="F7F9FC"; WHITE="FFFFFF"; GRAY="64748B"; GREEN="008000"; RED="C00000"; ORANGE="F59E0B"; LINE="DCE4EE"
THIN = Side(style="thin", color=LINE)
SECTOR_LABELS={
    "SHIPBUILDING":"조선",
    "DEFENSE":"방산",
    "POWER_EQUIPMENT":"전력기기",
    "CONSTRUCTION_EQUIPMENT":"건설장비",
    "MACHINERY":"기계",
}
SECTOR_ORDER={k:i for i,k in enumerate(SECTOR_LABELS)}


def _is_kr(r: dict) -> bool:
    return str(r.get("country") or "").upper()=="KR"


def _status_label(s):
    return "Coverage" if s=="COVERAGE" else ("NR" if s=="NR" else "미정")


def _country_label(v):
    return "대한민국" if str(v or "").upper()=="KR" else (v or "-")


def _data_status_label(v):
    labels={
        "COMPLETED_DAILY_QUOTE":"최신 완료 거래일",
        "COMPLETED_SESSION_SNAPSHOT":"최신 완료 거래일",
        "REFRESHED_COMPLETED_SESSION":"최신 완료 거래일",
        "COMPLETED_HISTORICAL_FALLBACK":"최신 완료 거래일",
        "COMPLETED_NO_COMPARISON_REFERENCE":"신규상장·직전 기준값 없음",
        "PRESERVED_OPEN_OR_UNKNOWN":"이전 완료 거래일 유지",
        "PRESERVED_AFTER_FETCH_ERROR":"수집 오류·이전값 유지",
        "FETCH_ERROR":"수집 오류",
    }
    return labels.get(v,v or "-")


def _sorted(rows):
    return sorted(rows,key=lambda r:(SECTOR_ORDER.get(r.get("research_sector"),99),0 if _is_kr(r) else 1,-float(r.get("market_cap") or 0),str(r.get("company_name") or "")))


def _title(ws, last_col, title, subtitle):
    ws.sheet_view.showGridLines=False
    ws.row_dimensions[1].height=31
    ws.row_dimensions[2].height=19
    ws["A1"]=title
    ws["A1"].font=Font(name="Arial",size=19,bold=True,color=WHITE)
    ws["A1"].fill=PatternFill("solid",fgColor=NAVY)
    ws["A1"].alignment=Alignment(vertical="center")
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=last_col)
    ws["A2"]=subtitle
    ws["A2"].font=Font(name="Arial",size=9,color="D6E4F5")
    ws["A2"].fill=PatternFill("solid",fgColor=NAVY)
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=last_col)


def _header(cell, text):
    cell.value=text
    cell.font=Font(name="Arial",size=9,bold=True,color=WHITE)
    cell.fill=PatternFill("solid",fgColor=NAVY2)
    cell.alignment=Alignment(horizontal="center",vertical="center")
    cell.border=Border(bottom=Side(style="thin",color="FFFFFF"))


def _make_table_sheet(ws, rows, qa=None, title="섹터별 주가 모니터", subtitle=""):
    headers=["섹터","구분","국가","회사","Ticker","거래소","리서치 상태","종가","전일종가/기준가","등락","등락률","TP","Upside","가격 거래일","데이터 상태","직전거래일","확인시각","가격 출처"]
    _title(ws,len(headers),title,subtitle)
    ws.row_dimensions[3].height=8
    for c,h in enumerate(headers,1): _header(ws.cell(4,c),h)
    ws.cell(4,1).comment=Comment("국내 Universe 분류 원천: NAVER Finance 업종분류 (https://finance.naver.com). 해외 Universe 분류 원천: TradingView Screener (https://www.tradingview.com).","OpenAI")
    ws.cell(4,14).comment=Comment("가격이 실제로 속한 거래일입니다. 해외는 TradingView 일봉 거래일 필드가 제공될 때 개별 종목별로 기록합니다.","OpenAI")
    ws.cell(4,15).comment=Comment("최신 완료 거래일인지, 장중/상태 불명으로 이전 완료 거래일을 유지했는지, 신규상장으로 직전 비교기준이 없는지 표시합니다.","OpenAI")
    for i,r in enumerate(_sorted(rows),5):
        sector=SECTOR_LABELS.get(r.get("research_sector"),r.get("research_sector") or "-")
        vals=[
            sector,"국내" if _is_kr(r) else "해외",_country_label(r.get("country")),r.get("company_name"),r.get("ticker"),r.get("exchange"),_status_label(r.get("research_status")),
            r.get("price"),r.get("previous_close"),r.get("price_change"),(r.get("price_change_pct")/100 if r.get("price_change_pct") is not None else None),r.get("target_price"),None,
            r.get("price_date") or "날짜 미확인",_data_status_label(r.get("data_status")),r.get("previous_trading_date"),r.get("last_checked_at") or r.get("price_observed_at"),r.get("price_source")
        ]
        for c,v in enumerate(vals,1):
            cell=ws.cell(i,c,v)
            cell.font=Font(name="Arial",size=9,color="000000")
            cell.alignment=Alignment(vertical="center",horizontal="right" if c in (8,9,10,11,12,13) else "left")
            if i%2==0: cell.fill=PatternFill("solid",fgColor=PALE)
        if r.get("target_price") and r.get("price"):
            ws.cell(i,13,f"=IFERROR(L{i}/H{i}-1,\"\")")
        for c in (8,9,10,12): ws.cell(i,c).number_format='#,##0.###;[Blue](#,##0.###);-'
        for c in (11,13): ws.cell(i,c).number_format='0.0%;[Blue](0.0%);-'
        if r.get("research_status")=="COVERAGE":
            ws.cell(i,7).fill=PatternFill("solid",fgColor="E8F0FF")
            ws.cell(i,7).font=Font(name="Arial",size=9,bold=True,color="214FB7")
        elif r.get("research_status")=="NR":
            ws.cell(i,7).fill=PatternFill("solid",fgColor="EEF0F3")
        if r.get("research_sector") in SECTOR_LABELS:
            ws.cell(i,1).font=Font(name="Arial",size=9,bold=True,color=NAVY2)
        if r.get("data_status") in ("PRESERVED_OPEN_OR_UNKNOWN","PRESERVED_AFTER_FETCH_ERROR","FETCH_ERROR","COMPLETED_NO_COMPARISON_REFERENCE"):
            ws.cell(i,15).fill=PatternFill("solid",fgColor="FFF4D8")
            ws.cell(i,15).font=Font(name="Arial",size=9,bold=True,color="A45B00")
    widths=[12,8,14,38,13,11,13,14,16,14,11,14,11,14,24,14,24,28]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="H5"
    ws.auto_filter.ref=f"A4:R{max(5,4+len(rows))}"
    ws.sheet_view.zoomScale=88
    if rows:
        last=4+len(rows)
        for col in ("J","K","M"):
            ws.conditional_formatting.add(f"{col}5:{col}{last}",CellIsRule(operator="greaterThan",formula=["0"],font=Font(color=RED,bold=True)))
            ws.conditional_formatting.add(f"{col}5:{col}{last}",CellIsRule(operator="lessThan",formula=["0"],font=Font(color=BLUE,bold=True)))


def build(rows: list[dict], qa: dict, path: str):
    rows=_sorted(rows)
    domestic=[r for r in rows if _is_kr(r)]
    global_rows=[r for r in rows if not _is_kr(r)]
    wb=Workbook()
    ws=wb.active;ws.title="Dashboard"
    subtitle=f"QA {qa['status']} | 전체 {len(rows):,}개 | 국내 {len(domestic):,}개 | 해외 {len(global_rows):,}개 | 생성 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    _make_table_sheet(ws,rows,qa,"섹터별 주가 모니터",subtitle)

    kd=wb.create_sheet("국내")
    _make_table_sheet(kd,domestic,qa,"국내 주가 모니터",f"NAVER/KRX 미러 완료 거래일 | {len(domestic):,}개")
    kg=wb.create_sheet("해외")
    _make_table_sheet(kg,global_rows,qa,"해외 주가 모니터",f"TradingView 완료 거래일 스냅샷 | {len(global_rows):,}개")

    u=wb.create_sheet("Universe");u.sheet_view.showGridLines=False
    uh=["company_name","ticker","country","exchange","currency","source","source_sector","source_industry","research_sector","research_status","target_price","target_currency","last_report_date","active","review_note"]
    for c,h in enumerate(uh,1): _header(u.cell(1,c),h)
    for i,r in enumerate(rows,2):
        for c,h in enumerate(uh,1):
            cell=u.cell(i,c,r.get(h,""));cell.font=Font(name="Arial",size=9)
            if i%2==0: cell.fill=PatternFill("solid",fgColor=PALE)
    u.freeze_panes="A2";u.auto_filter.ref=f"A1:O{max(2,1+len(rows))}";u.sheet_view.zoomScale=85
    for c in range(1,16):u.column_dimensions[get_column_letter(c)].width=18
    u.column_dimensions["A"].width=38;u.column_dimensions["H"].width=34;u.column_dimensions["O"].width=54

    q=wb.create_sheet("QA");q.sheet_view.showGridLines=False
    q["A1"]="QA STATUS";q["B1"]=qa["status"]
    for cell in (q["A1"],q["B1"]):cell.font=Font(bold=True,color=WHITE);cell.fill=PatternFill("solid",fgColor=NAVY)
    q["B1"].fill=PatternFill("solid",fgColor=RED if qa["status"]=="FAIL" else (ORANGE if qa["status"]=="REVIEW" else GREEN))
    summary=[
        ("전체 종목",qa.get("row_count")),("국내 종목",qa.get("domestic_count")),("가격 누락",qa.get("missing_price_count")),("수집 오류",qa.get("fetch_error_count")),
        ("국내 등락률 원천 확인",qa.get("official_kr_return_count")),("국내 최신 가격 거래일",qa.get("kr_latest_price_date")),("국내 최신 거래일 종목",qa.get("kr_latest_price_date_count")),
        ("국내 0% 종목",qa.get("kr_zero_return_count")),("국내 완료일 초과",qa.get("kr_future_date_count")),("직전 비교기준 없음",qa.get("missing_return_reference_count")),
        ("해외 거래일 확인",qa.get("global_price_date_count")),("해외 거래일 미확인",qa.get("global_price_date_missing_count")),
        ("해외 동일 거래소 대비 지연 후보",qa.get("global_lagging_price_date_count")),("해외 7일 이상 지연 후보",qa.get("global_severe_lagging_price_date_count")),
        ("해외 최신 완료 거래일 갱신",qa.get("refreshed_completed_global_count")),("해외 이전 완료 거래일 유지",qa.get("preserved_open_or_unknown_global_count")),
    ]
    row=3
    for label,value in summary:
        q.cell(row,1,label);q.cell(row,2,value);row+=1
    row+=1
    q.cell(row,1,"해외 거래일 분포");q.cell(row,1).font=Font(bold=True);row+=1
    for d,n in (qa.get("global_price_date_distribution") or {}).items():
        q.cell(row,1,d);q.cell(row,2,n);row+=1
    row+=1
    for level,items in (("ERROR",qa.get("errors",[])),("WARNING",qa.get("warnings",[]))):
        for msg in items:
            q.cell(row,1,level);q.cell(row,2,msg)
            if level=="ERROR": q.cell(row,1).fill=PatternFill("solid",fgColor="FDECEC")
            else: q.cell(row,1).fill=PatternFill("solid",fgColor="FFF4D8")
            row+=1
    q.column_dimensions["A"].width=34;q.column_dimensions["B"].width=110

    p=Path(path);p.parent.mkdir(parents=True,exist_ok=True);wb.save(p)
